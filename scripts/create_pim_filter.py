#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PIM Filter-Eigenschaften Generator

Dieses Skript liest die Masterliste (MasterlisteHC5.xlsx) ein und generiert
eine Excel-Datei mit:
- Sheet 1: Filtereigenschaften (Empfehlungen für PIM-Filter)
- Sheet 2: Artikelzuordnung (alle Artikel mit extrahierten Eigenschaften)
- Sheet 3: Wertelisten (alle einzigartigen Werte pro Eigenschaft)
- Sheet 4: Zusammenfassung (Statistiken)
- Sheet 5: Ableitungsregeln (Dokumentation der Ableitungslogik)

Abgeleitete Eigenschaften:
- Saison: aus Material + Produkttyp
- Muster: aus Farbnamen
- Anlass: aus Produkttyp + Beschreibung

Verwendung:
    python create_pim_filter.py

Voraussetzungen:
    - pandas
    - openpyxl
    - MasterlisteHC5.xlsx im gleichen Verzeichnis oder Pfad anpassen

Autor: Claude Code
Datum: 2026-01-28
"""

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Konfiguration - Pfade anpassen falls nötig
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # Parent directory (Projekt-Root)
INPUT_FILE = os.path.join(BASE_DIR, 'MasterlisteHC5.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, 'PIM_Filter_Eigenschaften.xlsx')

# Excel-Datei lesen
print(f"Lese Eingabedatei: {INPUT_FILE}")
df = pd.read_excel(INPUT_FILE)

# Spalten-Mapping
col_code = df.columns[2]
col_kopfartikel = df.columns[3]
col_beschreibung = df.columns[4]
col_kategorie = df.columns[9]
col_zusatz_kategorie = df.columns[10]
col_qualitaet = df.columns[18]
col_rohware = df.columns[19]
col_fertigung = df.columns[22]
col_zertifizierung = df.columns[23]
col_groessenlauf = df.columns[24]
col_pflege = df.columns[25]
col_print_lang = df.columns[13]
col_farbe1 = df.columns[27]
col_farbe2 = df.columns[28]
col_farbe3 = df.columns[29]
col_farbe4 = df.columns[30]
col_farbe5 = df.columns[31]

# Hilfsfunktionen
def extract_schnitt(text):
    if pd.isna(text):
        return None
    matches = re.findall(r'Schnitt:\s*([^\.]+)', str(text))
    if matches:
        return matches[0].strip()
    return None

def extract_passform(schnitt):
    if pd.isna(schnitt) or not schnitt:
        return None
    parts = str(schnitt).split('/')
    return parts[0].strip() if parts else None

def extract_laenge(schnitt):
    if pd.isna(schnitt) or not schnitt:
        return None
    laengen_map = {
        'mini': 'Mini',
        'kniefrei': 'Kniefrei',
        'knieumspielend': 'Knieumspielend',
        'kniebedeckt': 'Kniebedeckt',
        'knöchellang': 'Knöchellang',
        'sehr lang': 'Sehr lang',
        '7/8-länge': '7/8-Länge',
        'midi': 'Midi',
        'verkürzt': 'Verkürzt',
        'überknie': 'Überknie'
    }
    s = str(schnitt).lower()
    # Check for 'sehr lang' first (before 'lang')
    if 'sehr lang' in s:
        return 'Sehr lang'
    for key, val in laengen_map.items():
        if key in s:
            return val
    # Check 'lang' last to avoid matching 'knöchellang'
    if 'lang' in s and 'knöchel' not in s:
        return 'Lang'
    return None

def extract_aermellaenge(beschreibung, kategorie):
    if pd.isna(beschreibung):
        beschreibung = ''
    if pd.isna(kategorie):
        kategorie = ''
    text = str(beschreibung).lower() + ' ' + str(kategorie).lower()
    if '3/4-arm' in text or '3/4 arm' in text:
        return '3/4-Arm'
    elif 'langarm' in text:
        return 'Langarm'
    elif 'halbarm' in text:
        return 'Halbarm'
    elif 'kurzarm' in text:
        return 'Kurzarm'
    elif 'ärmellos' in text.lower() or 'top' == kategorie.lower().strip() or 'bra-top' == kategorie.lower().strip() or 'unterkleid' in text.lower():
        return 'Ärmellos'
    return None

def get_groessentyp(groessenlauf):
    if pd.isna(groessenlauf):
        return None
    g = str(groessenlauf)
    if 'XS' in g or 'XL' in g or 'S-L' in g or 'XXS' in g:
        return 'T-Shirt-Größen'
    elif re.match(r'^\d{2}-\d{2}$', g.strip()):
        return 'Konfektionsgrößen'
    elif 'A' in g or 'B' in g or 'C' in g:
        return 'BH-Größen'
    elif 'one size' in g.lower():
        return 'Einheitsgröße'
    return 'Sonstige'

def extract_saison(material, produkttyp):
    """Leitet die Saison aus Material und Produkttyp ab"""
    if pd.isna(material):
        material = ''
    if pd.isna(produkttyp):
        produkttyp = ''

    mat = str(material).lower()
    typ = str(produkttyp).lower()

    # Winter-Indikatoren (nur Yak und Schurwolle, NICHT 'wolle' wegen Baumwolle!)
    is_winter_mat = 'yak' in mat or 'schurwolle' in mat

    # Winter-Produkttypen
    winter_types = ['pullover', 'cardigan', 'strickshirt', 'pullunder', 'hoody', 'sweatshirt', 'sweatjacke', 'kapuzenjacke']
    is_winter_typ = any(w in typ for w in winter_types)

    # Sommer-Indikatoren: Leinen mit hohem Anteil (>=50%)
    is_sommer_mat = 'leinen' in mat and any(p in mat for p in ['50 %', '67 %', '68 %', '100 %'])

    # Sommer-Produkttypen
    sommer_types = ['top', 'bra-top', 'shorts', 'sandale']
    is_sommer_typ = any(s == typ.strip() or typ.startswith(s) for s in sommer_types)

    # Entscheidungslogik
    if is_winter_mat:
        return 'Winter'
    elif is_sommer_mat:
        return 'Sommer'
    elif is_winter_typ:
        return 'Winter/Übergang'
    elif is_sommer_typ:
        return 'Sommer'
    else:
        return 'Ganzjährig'

def extract_muster(farben_list):
    """Prüft ob Muster vorhanden ist anhand der Farbnamen"""
    muster_keywords = ['ringel', 'streif', 'karo', 'druck', 'print', 'batik', 'tupfen', 'punkte',
                       'blume', 'multicolor', 'multi', 'intarsia', 'ajour']
    melange_keywords = ['melange', 'mouliné', 'mix']

    for farbe in farben_list:
        if pd.isna(farbe):
            continue
        f = str(farbe).lower()
        # Echte Muster
        if any(k in f for k in muster_keywords):
            if 'ringel' in f:
                return 'Ringel'
            elif 'streif' in f:
                return 'Streifen'
            elif 'karo' in f:
                return 'Karo'
            elif any(k in f for k in ['druck', 'print', 'batik']):
                return 'Druck/Print'
            elif any(k in f for k in ['tupfen', 'punkte']):
                return 'Punkte'
            elif 'blume' in f:
                return 'Blumen'
            elif any(k in f for k in ['multi', 'intarsia']):
                return 'Mehrfarbig'
            elif 'ajour' in f:
                return 'Ajour'
        # Melange (meliert)
        if any(k in f for k in melange_keywords):
            return 'Melange'

    return 'Uni'

def extract_anlass(produkttyp, beschreibung):
    """Leitet den Anlass aus Produkttyp und Beschreibung ab"""
    if pd.isna(produkttyp):
        produkttyp = ''
    if pd.isna(beschreibung):
        beschreibung = ''

    typ = str(produkttyp).lower().strip()
    desc = str(beschreibung).lower()
    combined = typ + ' ' + desc

    # Sport/Yoga - nur wenn explizit Yoga/Sport im Namen
    if 'yoga' in combined or 'sport' in combined:
        return 'Sport/Yoga'

    # Home/Schlafen
    if any(k in typ for k in ['nachthemd', 'pyjama', 'schlaf', 'nachtwäsche']):
        return 'Home/Schlafen'

    # Büro/Smart Casual
    if any(k in typ for k in ['blusenshirt', 'hemdbluse', 'bluse']):
        return 'Büro/Smart Casual'

    # Wäsche
    if any(k in typ for k in ['bh', 'slip', 'hipster', 'unterhemd', 'unterkleid']):
        return 'Unterwäsche'

    # Schuhe
    if any(k in typ for k in ['sandale', 'sneaker', 'slipper']):
        return 'Vielseitig'

    # Accessoires
    if any(k in typ for k in ['gürtel', 'bandana']):
        return 'Accessoire'

    # Default: Casual/Freizeit (inkl. Tops, T-Shirts, Hosen, etc.)
    return 'Casual/Freizeit'

# Workbook erstellen
wb = Workbook()

# ======= SHEET 1: Filtereigenschaften =======
ws1 = wb.active
ws1.title = "Filtereigenschaften"

# Header-Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Filtereigenschaften definieren
filter_properties = [
    ["Filtername", "Beschreibung", "Filtertyp", "Werte aus Masterliste", "PIM-Status", "Empfehlung"],

    # Kategorie-Filter
    ["Hauptkategorie", "Primäre Produktkategorie", "Single-Select",
     "Oberteile, Unterteile, Kleid, Overall, Nachtwäsche, Wäsche, Accessoires, Schuhe",
     "Vorhanden (Kategorie)", "Pflicht-Filter"],

    ["Produkttyp", "Detaillierte Produktart", "Single-Select",
     "Top, Bra-Top, Langarm-Shirt, Kurzarm-Shirt, Blusenshirt, Pullover, Cardigan, Jacke, Hoody, Sweatshirt, Leggings, Hose, Culotte, Rock, Kleid, Overall, BH, Slip, Hipster, Nachthemd, Pyjama, Sandale, Sneaker, Slipper, Bandana, Gürtel",
     "Vorhanden (Zusätzliche Kategorie)", "Pflicht-Filter"],

    # Passform-Filter
    ["Passform", "Wie das Kleidungsstück sitzt", "Single-Select",
     "figurbetont, figurumspielend, leger, sehr leger, oversize, weit, schmal, skinny, gerade, A-Linie, Kimonoform, leichte A-Linie",
     "Nicht vorhanden", "Neuer Filter - Kundenrelevant"],

    # Länge-Filter
    ["Länge", "Länge bei Hosen/Röcken/Kleidern", "Single-Select",
     "Mini, Kniefrei, Knieumspielend, Kniebedeckt, Midi, Knöchellang, Lang, Sehr lang, 7/8-Länge, Verkürzt, Überknie",
     "Nicht vorhanden", "Neuer Filter - Kundenrelevant"],

    # Ärmellänge
    ["Ärmellänge", "Länge der Ärmel", "Single-Select",
     "Ärmellos, Kurzarm, Halbarm, 3/4-Arm, Langarm",
     "Vorhanden (Ärmellänge)", "Vorhandener Filter erweitern"],

    # Größen-Filter
    ["Größentyp", "Typ des Größensystems", "Multi-Select",
     "T-Shirt-Größen (XS-XL), Konfektionsgrößen (32-46), BH-Größen (75A-90C), Schuhgrößen (36-41), Einheitsgröße",
     "Vorhanden (Konfektionsgröße, Größe)", "Getrennte Größengruppen anlegen"],

    # Material-Filter
    ["Material-Hauptkomponente", "Hauptbestandteil des Materials", "Multi-Select",
     "Baumwolle kbA, Leinen kbA, Yak, Schurwolle, Hanf kbA, Leder",
     "Vorhanden (Material)", "Filter optimieren"],

    ["Material-Qualität", "Vollständige Materialzusammensetzung", "Text/Info",
     "100% Baumwolle kbA, 100% Leinen kbA, 50% Yak/50% Schurwolle, diverse Mischgewebe",
     "Vorhanden (Material)", "Als Produktinfo anzeigen"],

    # Pflege-Filter
    ["Pflegehinweis", "Waschtemperatur/Pflegeart", "Single-Select",
     "30°, 40°, Handwäsche",
     "Nicht vorhanden", "Neuer Filter - Optional"],

    # Nachhaltigkeits-Info (keine Filter, nur Produktinfo)
    ["Zertifizierung", "Bio-/Öko-Zertifikat", "Info (kein Filter)",
     "organic (GOTS), made with organic, IVN Naturleder",
     "Vorhanden (Zertifzierungen)", "Nur Produktinfo"],

    ["Herkunft Rohware", "Ursprungsland der Materialien", "Info (kein Filter)",
     "Türkei, Indien, Frankreich, Belgien, Mongolei, Deutschland, Tadschikistan, Tansania, Israel, Litauen, Griechenland, Spanien",
     "Nicht vorhanden", "Nur Produktinfo"],

    ["Produktionsland", "Land der Fertigung", "Info (kein Filter)",
     "Portugal, Litauen, Türkei, Bulgarien, Mongolei, Kroatien, Rumänien",
     "Nicht vorhanden", "Nur Produktinfo"],

    # Neue abgeleitete Filter
    ["Saison", "Jahreszeit für das Kleidungsstück", "Single-Select",
     "Sommer, Winter, Winter/Übergang, Ganzjährig",
     "Nicht vorhanden", "Neuer Filter - Abgeleitet aus Material/Produkttyp"],

    ["Muster", "Art des Musters/Designs", "Single-Select",
     "Uni, Ringel, Streifen, Karo, Druck/Print, Punkte, Blumen, Mehrfarbig, Melange, Ajour",
     "Nicht vorhanden", "Neuer Filter - Abgeleitet aus Farbnamen"],

    ["Anlass", "Verwendungszweck/Gelegenheit", "Single-Select",
     "Casual/Freizeit, Sport/Yoga, Büro/Smart Casual, Home/Schlafen, Unterwäsche, Vielseitig, Accessoire",
     "Nicht vorhanden", "Neuer Filter - Abgeleitet aus Produkttyp"],
]

# Daten schreiben
for row_idx, row_data in enumerate(filter_properties, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Spaltenbreiten
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 70
ws1.column_dimensions['E'].width = 35
ws1.column_dimensions['F'].width = 30

# ======= SHEET 2: Artikelzuordnung =======
ws2 = wb.create_sheet("Artikelzuordnung")

# Artikeldaten sammeln
artikel_data = []
for idx, row in df.iterrows():
    kopfartikel = row[col_kopfartikel]

    if pd.isna(kopfartikel):
        continue

    schnitt = extract_schnitt(row[col_print_lang])

    # Farben für Muster-Erkennung sammeln
    farben = [row[col_farbe1], row[col_farbe2], row[col_farbe3], row[col_farbe4], row[col_farbe5]]

    artikel_data.append({
        'Artikelnummer': kopfartikel,
        'Code': row[col_code] if not pd.isna(row[col_code]) else '',
        'Hauptkategorie': row[col_kategorie] if not pd.isna(row[col_kategorie]) else '',
        'Produkttyp': row[col_zusatz_kategorie] if not pd.isna(row[col_zusatz_kategorie]) else '',
        'Material': row[col_qualitaet] if not pd.isna(row[col_qualitaet]) else '',
        'Passform': extract_passform(schnitt) if schnitt else '',
        'Länge': extract_laenge(schnitt) if schnitt else '',
        'Ärmellänge': extract_aermellaenge(row[col_beschreibung], row[col_zusatz_kategorie]) if extract_aermellaenge(row[col_beschreibung], row[col_zusatz_kategorie]) else '',
        'Größentyp': get_groessentyp(row[col_groessenlauf]) if get_groessentyp(row[col_groessenlauf]) else '',
        'Größenlauf': row[col_groessenlauf] if not pd.isna(row[col_groessenlauf]) else '',
        'Pflegehinweis': row[col_pflege] if not pd.isna(row[col_pflege]) else '',
        'Saison': extract_saison(row[col_qualitaet], row[col_zusatz_kategorie]),
        'Muster': extract_muster(farben),
        'Anlass': extract_anlass(row[col_zusatz_kategorie], row[col_beschreibung]),
        'Zertifizierung': row[col_zertifizierung] if not pd.isna(row[col_zertifizierung]) else '',
        'Herkunft_Rohware': row[col_rohware] if not pd.isna(row[col_rohware]) else '',
        'Produktionsland': row[col_fertigung] if not pd.isna(row[col_fertigung]) else ''
    })

# Header schreiben
headers = ['Artikelnummer', 'Code', 'Hauptkategorie', 'Produkttyp', 'Material', 'Passform',
           'Länge', 'Ärmellänge', 'Größentyp', 'Größenlauf', 'Pflegehinweis',
           'Saison', 'Muster', 'Anlass',
           'Zertifizierung', 'Herkunft_Rohware', 'Produktionsland']

for col_idx, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

# Daten schreiben
for row_idx, artikel in enumerate(artikel_data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=artikel.get(header, ''))
        cell.border = thin_border

# Spaltenbreiten für Sheet 2 (erweitert um Saison, Muster, Anlass)
col_widths = {'A': 15, 'B': 12, 'C': 15, 'D': 22, 'E': 45, 'F': 18,
              'G': 15, 'H': 12, 'I': 20, 'J': 12, 'K': 12,
              'L': 15, 'M': 15, 'N': 18,  # Saison, Muster, Anlass
              'O': 15, 'P': 35, 'Q': 18}  # Zertifizierung, Herkunft, Produktionsland
for col, width in col_widths.items():
    ws2.column_dimensions[col].width = width

# ======= SHEET 3: Wertelisten =======
ws3 = wb.create_sheet("Wertelisten")

# Alle einzigartigen Werte sammeln
unique_kategorien = sorted([str(x) for x in df[col_kategorie].dropna().unique()])
unique_produkttypen = sorted([str(x) for x in df[col_zusatz_kategorie].dropna().unique()])
unique_passformen = set()
unique_laengen = set()
unique_aermellaengen = set()
unique_materialien = sorted([str(x) for x in df[col_qualitaet].dropna().unique()])
unique_pflege = sorted([str(x) for x in df[col_pflege].dropna().unique()])
unique_zertifizierung = sorted([str(x) for x in df[col_zertifizierung].dropna().unique()])
unique_groessentypen = set()
unique_herkunft = sorted([str(x) for x in df[col_rohware].dropna().unique()])
unique_produktion = sorted([str(x) for x in df[col_fertigung].dropna().unique()])
unique_saison = set()
unique_muster = set()
unique_anlass = set()

for artikel in artikel_data:
    if artikel['Passform']:
        unique_passformen.add(artikel['Passform'])
    if artikel['Länge']:
        unique_laengen.add(artikel['Länge'])
    if artikel['Ärmellänge']:
        unique_aermellaengen.add(artikel['Ärmellänge'])
    if artikel['Größentyp']:
        unique_groessentypen.add(artikel['Größentyp'])
    if artikel['Saison']:
        unique_saison.add(artikel['Saison'])
    if artikel['Muster']:
        unique_muster.add(artikel['Muster'])
    if artikel['Anlass']:
        unique_anlass.add(artikel['Anlass'])

# Wertelisten schreiben
wertelisten = [
    ("Hauptkategorie", unique_kategorien),
    ("Produkttyp", unique_produkttypen),
    ("Passform", sorted(unique_passformen)),
    ("Länge", sorted(unique_laengen)),
    ("Ärmellänge", sorted(unique_aermellaengen)),
    ("Größentyp", sorted(unique_groessentypen)),
    ("Material", unique_materialien),
    ("Pflegehinweis", unique_pflege),
    ("Saison", sorted(unique_saison)),
    ("Muster", sorted(unique_muster)),
    ("Anlass", sorted(unique_anlass)),
    ("Zertifizierung", unique_zertifizierung),
    ("Herkunft Rohware", unique_herkunft),
    ("Produktionsland", unique_produktion),
]

col_offset = 1
for liste_name, werte in wertelisten:
    # Header
    cell = ws3.cell(row=1, column=col_offset, value=liste_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

    # Werte
    for row_idx, wert in enumerate(werte, 2):
        cell = ws3.cell(row=row_idx, column=col_offset, value=wert)
        cell.border = thin_border

    # Spaltenbreite basierend auf Inhalt
    max_len = max([len(str(w)) for w in werte] + [len(liste_name)])
    ws3.column_dimensions[chr(64 + col_offset)].width = min(max_len + 2, 45)
    col_offset += 1

# ======= SHEET 4: Zusammenfassung =======
ws4 = wb.create_sheet("Zusammenfassung")

summary_data = [
    ["Zusammenfassung der PIM-Filtereigenschaften"],
    [""],
    ["Anzahl Artikel in Masterliste:", len(artikel_data)],
    [""],
    ["EMPFOHLENE NEUE FILTER (kundenrelevant):"],
    ["- Passform", f"{len(unique_passformen)} Werte"],
    ["- Länge", f"{len(unique_laengen)} Werte"],
    ["- Pflegehinweis", f"{len(unique_pflege)} Werte"],
    [""],
    ["ABGELEITETE FILTER (automatisch berechnet):"],
    ["- Saison", f"{len(unique_saison)} Werte (aus Material + Produkttyp)"],
    ["- Muster", f"{len(unique_muster)} Werte (aus Farbnamen)"],
    ["- Anlass", f"{len(unique_anlass)} Werte (aus Produkttyp)"],
    [""],
    ["VORHANDENE FILTER ZU ERWEITERN:"],
    ["- Ärmellänge", f"{len(unique_aermellaengen)} Werte"],
    ["- Größentyp (getrennte Gruppen)", f"{len(unique_groessentypen)} Gruppen"],
    [""],
    ["NUR ALS PRODUKTINFO (keine Filter):"],
    ["- Zertifizierung", f"{len(unique_zertifizierung)} Werte"],
    ["- Herkunft Rohware", f"{len(unique_herkunft)} Werte"],
    ["- Produktionsland", f"{len(unique_produktion)} Werte"],
    [""],
    ["STATISTIK NACH HAUPTKATEGORIE:"],
]

# Kategorie-Statistik
for kat in unique_kategorien:
    count = len([a for a in artikel_data if a['Hauptkategorie'] == kat])
    summary_data.append([f"  {kat}:", f"{count} Artikel"])

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif 'EMPFOHLEN' in str(value) or 'ABGELEITET' in str(value) or 'VORHANDEN' in str(value) or 'NUR ALS' in str(value) or 'STATISTIK' in str(value):
            cell.font = Font(bold=True)

ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 20

# ======= SHEET 5: Ableitungsregeln =======
ws5 = wb.create_sheet("Ableitungsregeln")

# Styles für Überschriften
section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=12)

regeln_data = [
    # SAISON
    ["SAISON - Ableitungsregeln", "", ""],
    ["Abgeleitet aus:", "Material (Qualität) + Produkttyp", ""],
    ["", "", ""],
    ["Wert", "Regel", "Beispiel"],
    ["Winter", "Material enthält 'yak' ODER 'schurwolle'", "50% Yak, 50% Schurwolle → Winter"],
    ["Sommer", "Material enthält 'leinen' mit Anteil ≥50% (50%, 67%, 68%, 100%)", "100% Leinen kbA → Sommer"],
    ["Sommer", "ODER Produkttyp ist 'top', 'bra-top', 'shorts', 'sandale'", "Top → Sommer"],
    ["Winter/Übergang", "Produkttyp ist Pullover, Cardigan, Strickshirt, Pullunder, Hoody, Sweatshirt, Sweatjacke, Kapuzenjacke (ohne Woll-Material)", "Pullover aus Baumwolle → Winter/Übergang"],
    ["Ganzjährig", "Alle anderen (reine Baumwolle, niedrige Leinen-Anteile, Hanf-Mix)", "100% Baumwolle kbA → Ganzjährig"],
    ["", "", ""],
    ["", "", ""],
    # MUSTER
    ["MUSTER - Ableitungsregeln", "", ""],
    ["Abgeleitet aus:", "Farbnamen (Spalten Farben 1-5)", ""],
    ["", "", ""],
    ["Wert", "Regel (Keyword im Farbnamen)", "Beispiel"],
    ["Ringel", "Farbname enthält 'ringel'", "blau/curry ringel → Ringel"],
    ["Streifen", "Farbname enthält 'streif'", "blau/weiss streif → Streifen"],
    ["Karo", "Farbname enthält 'karo'", "karo azurblau/multicolor → Karo"],
    ["Druck/Print", "Farbname enthält 'druck', 'print' oder 'batik'", "batikdruck pazifik → Druck/Print"],
    ["Punkte", "Farbname enthält 'tupfen' oder 'punkte'", "punkte druck → Punkte"],
    ["Blumen", "Farbname enthält 'blume'", "blume weiss → Blumen"],
    ["Mehrfarbig", "Farbname enthält 'multi' oder 'intarsia'", "multistreif blau → Mehrfarbig"],
    ["Ajour", "Farbname enthält 'ajour'", "smaragd ajour ringel → Ajour"],
    ["Melange", "Farbname enthält 'melange', 'mouliné' oder 'mix'", "hasel melange → Melange"],
    ["Uni", "Keine der obigen Keywords gefunden", "schwarz, navy, creme → Uni"],
    ["", "", ""],
    ["", "", ""],
    # ANLASS
    ["ANLASS - Ableitungsregeln", "", ""],
    ["Abgeleitet aus:", "Produkttyp + Beschreibung", ""],
    ["", "", ""],
    ["Wert", "Regel", "Beispiel"],
    ["Sport/Yoga", "Produkttyp oder Beschreibung enthält 'yoga' oder 'sport'", "Yogahose, Yoga-Top → Sport/Yoga"],
    ["Home/Schlafen", "Produkttyp enthält 'nachthemd', 'pyjama', 'schlaf' oder 'nachtwäsche'", "Pyjama-Hose → Home/Schlafen"],
    ["Büro/Smart Casual", "Produkttyp enthält 'blusenshirt', 'hemdbluse' oder 'bluse'", "Blusenshirt → Büro/Smart Casual"],
    ["Unterwäsche", "Produkttyp enthält 'bh', 'slip', 'hipster', 'unterhemd' oder 'unterkleid'", "BH, Hipster → Unterwäsche"],
    ["Vielseitig", "Produkttyp enthält 'sandale', 'sneaker' oder 'slipper'", "Sandale → Vielseitig"],
    ["Accessoire", "Produkttyp enthält 'gürtel' oder 'bandana'", "Gürtel → Accessoire"],
    ["Casual/Freizeit", "Alle anderen (T-Shirts, Hosen, Röcke, Kleider, etc.)", "Kurzarm-Shirt, Hose → Casual/Freizeit"],
]

# Header für Sektionen
section_rows = [1, 12, 28]  # Zeilen mit Sektions-Überschriften
header_rows = [4, 15, 31]   # Zeilen mit Spalten-Überschriften

for row_idx, row_data in enumerate(regeln_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

        # Sektions-Überschriften (grün)
        if row_idx in section_rows:
            cell.fill = section_fill
            cell.font = section_font
        # Spalten-Überschriften (blau)
        elif row_idx in header_rows:
            cell.fill = header_fill
            cell.font = header_font
        # "Abgeleitet aus" Zeilen
        elif 'Abgeleitet aus:' in str(value):
            cell.font = Font(italic=True)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 70
ws5.column_dimensions['C'].width = 45

# Speichern mit Fehlerbehandlung
save_path = OUTPUT_FILE
try:
    print(f"Speichere Ausgabedatei: {save_path}")
    wb.save(save_path)
except PermissionError:
    # Falls Datei geöffnet ist, mit Zeitstempel speichern
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_path = OUTPUT_FILE.replace('.xlsx', f'_{timestamp}.xlsx')
    print(f"Original-Datei ist geöffnet. Speichere als: {save_path}")
    wb.save(save_path)

print(f'Excel-Datei erstellt: {save_path}')
print(f'- Sheet 1: Filtereigenschaften ({len(filter_properties)-1} Eigenschaften)')
print(f'- Sheet 2: Artikelzuordnung ({len(artikel_data)} Artikel)')
print(f'- Sheet 3: Wertelisten ({len(wertelisten)} Listen)')
print(f'- Sheet 4: Zusammenfassung')
print(f'- Sheet 5: Ableitungsregeln (Saison, Muster, Anlass)')
